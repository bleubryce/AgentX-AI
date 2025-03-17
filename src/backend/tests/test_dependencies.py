import pytest
import pandas as pd
import openpyxl
import stripe
from io import BytesIO
from fastapi import FastAPI
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from weasyprint import HTML

def test_pandas_dependency():
    """Test if pandas is working correctly."""
    # Create a simple DataFrame
    df = pd.DataFrame({'A': [1, 2, 3], 'B': [4, 5, 6]})
    assert df.shape == (3, 2)
    assert list(df.columns) == ['A', 'B']

def test_openpyxl_dependency():
    """Test if openpyxl is working correctly."""
    # Create a workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'Test'
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Read back
    wb2 = openpyxl.load_workbook(buffer)
    ws2 = wb2.active
    assert ws2['A1'].value == 'Test'

def test_weasyprint_dependency():
    """Test if weasyprint is working correctly."""
    # Create a simple HTML document
    html = HTML(string='<h1>Test</h1>')
    buffer = BytesIO()
    html.write_pdf(buffer)
    assert buffer.getvalue()[:4] == b'%PDF'  # PDF signature

def test_stripe_dependency():
    """Test if stripe is properly imported."""
    # Just check if we can access the API version
    assert hasattr(stripe, 'api_version')
    assert isinstance(stripe.api_version, str)

def test_fastapi_dependency():
    """Test if FastAPI is working correctly."""
    app = FastAPI()
    
    @app.get("/test")
    def read_test():
        return {"message": "Test successful"}
    
    client = TestClient(app)
    response = client.get("/test")
    assert response.status_code == 200
    assert response.json() == {"message": "Test successful"}

def test_sqlalchemy_dependency():
    """Test if SQLAlchemy is working correctly."""
    # Create an in-memory SQLite engine
    engine = create_engine("sqlite:///:memory:")
    SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    
    # Just check if we can create a session
    session = SessionLocal()
    assert session is not None

if __name__ == "__main__":
    # Run tests manually
    test_pandas_dependency()
    test_openpyxl_dependency()
    test_weasyprint_dependency()
    test_stripe_dependency()
    test_fastapi_dependency()
    test_sqlalchemy_dependency()
    print("All backend dependency tests passed!") 